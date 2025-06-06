import random
from datetime import datetime
from openpyxl import load_workbook


def modify_excel_file(
    input_path: str,
    output_path: str,
    variation_percent: float = 10.0,   # ±10% random noise on each reading
    ev_profile_kw: float = 3.6,        # extra kW to add when EV is charging
    ev_hours: tuple = (18, 22)         # charge from 18:00 up to 21:59
) -> None:
    """
    Loads an existing .xlsx, leaves row 1 and row 2 entirely intact,
    then for each data row (row 3 onward):
      • Randomly perturbs the “Active power (kW)” cell by ±variation_percent
      • If that timestamp’s date is one of 50% randomly chosen days
        AND the hour falls between ev_hours[0] and ev_hours[1]-1,
        adds ev_profile_kw to the Active power value
      • Replaces the “month” column’s formula (=MONTH(Ax)) with an integer.
    Finally, saves to output_path, preserving ALL original formatting except
    that numeric cells have new values (and “month” formulas are now static).
    """

    # 1) load the workbook in “normal” mode (preserves fonts, widths, formulae, etc.)
    wb = load_workbook(input_path)
    ws = wb.active  # we know 'Foglio1' is the active sheet

    # 2) build a map of header‐name → column_index by reading row 2 exactly
    #    (we do NOT touch row 1 at all; it stays the same)
    header_map = {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(ws[2], start=1)
    }
    # expects: header_map["Date and time"] → e.g. 1
    #          header_map["Active power (kW)"] → e.g. 2
    #          header_map["month"] → e.g. 3
    date_col  = header_map.get("Date and time")
    power_col = header_map.get("Active power (kW)")
    month_col = header_map.get("month")

    if date_col is None or power_col is None or month_col is None:
        raise ValueError(
            "Row 2 must contain exactly these three headers: "
            "'Date and time', 'Active power (kW)', 'month'."
        )

    # 3) collect all unique calendar‐days from the “Date and time” column (row 3 onward)
    unique_days = set()
    for row in ws.iter_rows(min_row=3, min_col=date_col, max_col=date_col):
        cell = row[0]
        if isinstance(cell.value, datetime):
            unique_days.add(cell.value.date())

    # pick 50% of those days at random for EV charging
    ev_days = random.sample(sorted(unique_days), k=int(len(unique_days) * 0.5))

    # 4) Walk through each data row (row 3 → last row). Only edit two cells:
    #      • “Active power (kW)”  (with random ± variation + EV load if applicable)
    #      • “month”               (overwrite its formula with a plain integer)
    #    Everything else (row 1, cells in row 2, any other columns) stays exactly
    #    as it was on disk (styles, widths, comments, hidden rows—nothing is changed).
    for row in ws.iter_rows(min_row=3):
        date_cell  = row[date_col - 1]   # e.g. cell A3, A4, A5, ...
        power_cell = row[power_col - 1]  # e.g. cell B3, B4, B5, ...
        mon_cell   = row[month_col - 1]  # e.g. cell C3, C4, C5, ...

        # skip if “Date and time” is missing or not a datetime
        if not isinstance(date_cell.value, datetime):
            continue
        # skip if “Active power” is missing or not numeric
        if not isinstance(power_cell.value, (int, float)):
            continue

        base_power = power_cell.value
        # apply ±variation_percent% randomness
        factor = 1 + random.uniform(-variation_percent / 100,
                                     variation_percent / 100)
        new_power = base_power * factor

        # if this row’s date is in our random ev_days *and* its hour is in ev_hours,
        # add the EV charging load
        row_date = date_cell.value.date()
        row_hour = date_cell.value.hour
        if (row_date in ev_days) and (ev_hours[0] <= row_hour < ev_hours[1]):
            new_power += ev_profile_kw

        # STEP 4a: write “Active power (kW)” = round(new_power, 3),
        #          preserving the cell’s original number_format
        saved_num_fmt = power_cell.number_format
        power_cell.value = round(new_power, 3)
        power_cell.number_format = saved_num_fmt

        # STEP 4b: overwrite the “month” cell so it becomes a plain integer
        #          (month number) instead of a =MONTH(...) formula.
        #          We know date_cell.value is a datetime, so:
        mon_cell.value = row_date.month
        # (preserve whatever number_format was there—likely “General” or “0”)
        saved_mon_fmt = mon_cell.number_format
        mon_cell.number_format = saved_mon_fmt

    # 5) Save out the new workbook.  Everything except the numeric cells above is untouched.
    wb.save(output_path)


# === Generate three modified workbooks ===
SOURCE = "load profile.xlsx"
for i in range(1, 5):
    target = f"load_profile_{i}.xlsx"
    modify_excel_file(SOURCE, target)
